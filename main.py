from flask import Flask, render_template, request, url_for, jsonify, redirect
import os
from datetime import datetime
import json
import requests
import re

from flask import send_file
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from copy import copy

app = Flask(__name__)
UPLOAD_FOLDER = 'static/uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


@app.route('/generate', methods=['POST'])
def generate():
    urls_json = request.form.get('urls')
    try:
        urls = json.loads(urls_json)
    except:
        urls = []

    return render_template('generate.html', urls=urls)


def generate_filename(ext):
    now = datetime.now()
    date_part = now.strftime('%d%m%Y')
    midnight = now.replace(hour=0, minute=0, second=0, microsecond=0)
    delta_ms = int((now - midnight).total_seconds() * 1000)
    return f"{date_part}_{delta_ms}.{ext}"


@app.route('/', methods=['GET', 'POST'])
def index():
    image_urls = []

    if request.method == 'POST':
        files = request.files.getlist('images')
        descriptions = request.form.getlist('descriptions')
        for i, file in enumerate(files):
            if file.filename != '':
                ext = file.filename.split('.')[-1].lower()
                filename = generate_filename(ext)
                description = descriptions[i] if i < len(descriptions) else ""
                save_metadata(filename, description)

                # Đảm bảo không trùng tên
                while os.path.exists(os.path.join(UPLOAD_FOLDER, filename)):
                    filename = generate_filename(ext)

                save_path = os.path.join(UPLOAD_FOLDER, filename)
                file.save(save_path)
                metadata = load_metadata()
                image_urls.append({
                    'url': url_for('static', filename=f'uploads/{file}'),
                    'description': metadata.get(file, '')
                })

    # Load ảnh đã lưu
    def extract_datetime_from_filename(filename):
        try:
            name = filename.rsplit('.', 1)[0]  # Bỏ phần .png, .jpg,...
            date_part, ms_part = name.split('_')
            dt = datetime.strptime(date_part, "%d%m%Y")
            return dt.timestamp() * 1000 + int(ms_part)
        except Exception:
            return 0  # nếu lỗi định dạng thì cho xuống cuối

    files = [f for f in os.listdir(UPLOAD_FOLDER)
             if f.lower().endswith(('png', 'jpg', 'jpeg', 'gif', 'bmp'))]

    files.sort(key=extract_datetime_from_filename, reverse=True)

    metadata = load_metadata()
    for file in files:
        image_urls.append({
            'url': url_for('static', filename=f'uploads/{file}'),
            'description': metadata.get(file, '')
        })

    return render_template('index.html', image_urls=image_urls)


@app.route('/delete', methods=['DELETE'])
def delete_image():
    filename = request.args.get('filename')
    if not filename:
        return jsonify({"error": "Missing filename"}), 400

    file_path = os.path.join(UPLOAD_FOLDER, filename)
    if os.path.exists(file_path):
        os.remove(file_path)
        return jsonify({"status": "success"})
    else:
        return jsonify({"error": "File not found"}), 404


@app.route("/generate-test-script", methods=["POST"])
def generate_test_script():
    data = request.get_json()

    try:
        res = requests.post(
            # "http://localhost:5000/get-test-scenario",
            "http://host.docker.internal:5000/get-test-scenario",
            json={
                "images": data.get("images", []),
                "action": data.get("action", "")
            },
            headers={"Content-Type": "application/json"}
        )

        res_json = res.json()
        content_list = res_json.get('content', [])

        return "\n".join(content_list), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


def save_metadata(filename, description):
    metadata_path = os.path.join(UPLOAD_FOLDER, 'metadata.json')
    try:
        with open(metadata_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except:
        data = {}

    data[filename] = description
    with open(metadata_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_metadata():
    metadata_path = os.path.join(UPLOAD_FOLDER, 'metadata.json')
    try:
        with open(metadata_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        return {}


@app.route('/update-description', methods=['POST'])
def update_description():
    data = request.get_json()
    filename = data.get('filename')
    new_description = data.get('description')

    if not filename:
        return jsonify({'error': 'Thiếu filename'}), 400

    metadata_path = os.path.join(UPLOAD_FOLDER, 'metadata.json')
    try:
        with open(metadata_path, 'r', encoding='utf-8') as f:
            metadata = json.load(f)
    except:
        metadata = {}

    if filename not in metadata:
        return jsonify({'error': 'Ảnh không tồn tại trong metadata'}), 404

    metadata[filename] = new_description
    with open(metadata_path, 'w', encoding='utf-8') as f:
        json.dump(metadata, f, ensure_ascii=False, indent=2)

    return jsonify({'status': 'success'})


@app.route("/save-metadata", methods=["POST"])
def save_metadata_api():
    data = request.get_json()

    image_filenames = [url.split("/")[-1] for url in data["images"]]

    dataset = {
        "images": image_filenames,
        "feature_desc": data["feature_desc"],
        "generated_test_script": data["generated_test_script"],
        "ground_truth_script": data["ground_truth_script"]
    }

    metadata_path = "static/uploads/metadata_submission.json"

    # Load existing list or create new
    if os.path.exists(metadata_path):
        with open(metadata_path, "r") as f:
            try:
                metadata_list = json.load(f)
                if not isinstance(metadata_list, list):
                    metadata_list = []
            except json.JSONDecodeError:
                metadata_list = []
    else:
        metadata_list = []

    # Append new dataset
    metadata_list.append(dataset)

    # Save back
    with open(metadata_path, "w") as f:
        json.dump(metadata_list, f, indent=2)

    return jsonify({"status": "ok"})


@app.route("/view-dataset")
def view_dataset():
    return render_template("view_dataset.html")


@app.route("/get-dataset")
def get_dataset():
    metadata_path = "static/uploads/metadata_submission.json"
    if os.path.exists(metadata_path):
        with open(metadata_path, "r") as f:
            try:
                data = json.load(f)
                return jsonify(data)
            except json.JSONDecodeError:
                return jsonify([])
    return jsonify([])


@app.route("/update-dataset", methods=["POST"])
def update_dataset():
    payload = request.get_json()
    index = payload.get("index")
    item = payload.get("item")

    metadata_path = "static/uploads/metadata_submission.json"
    if os.path.exists(metadata_path):
        with open(metadata_path, "r") as f:
            data = json.load(f)
    else:
        data = []

    if 0 <= index < len(data):
        data[index] = item
        with open(metadata_path, "w") as f:
            json.dump(data, f, indent=2)

    return jsonify({"status": "ok"})


@app.route("/delete-dataset", methods=["POST"])
def delete_dataset():
    payload = request.get_json()
    index = payload.get("index")

    metadata_path = "static/uploads/metadata_submission.json"
    if os.path.exists(metadata_path):
        with open(metadata_path, "r") as f:
            data = json.load(f)
    else:
        data = []

    if 0 <= index < len(data):
        data.pop(index)
        with open(metadata_path, "w") as f:
            json.dump(data, f, indent=2)

    return jsonify({"status": "ok"})


@app.route("/regenerate/<int:index>")
def regenerate(index):
    metadata_path = "static/uploads/metadata_submission.json"
    dataset = None

    if os.path.exists(metadata_path):
        with open(metadata_path, "r") as f:
            try:
                data = json.load(f)
                if 0 <= index < len(data):
                    dataset = data[index]
            except json.JSONDecodeError:
                dataset = None

    if not dataset:
        return redirect("/view-dataset")

    # convert file list thành URL
    urls = [url_for('static', filename=f'uploads/{img}') for img in dataset["images"]]

    return render_template(
        "regenerate.html",
        dataset_index=index,
        urls=urls,
        feature_desc=dataset.get("feature_desc", ""),
        generated_test_script=dataset.get("generated_test_script", ""),
        ground_truth_script=dataset.get("ground_truth_script", "")
    )


@app.route("/update-dataset-item/<int:index>", methods=["POST"])
def update_dataset_item(index):
    payload = request.get_json()
    metadata_path = "static/uploads/metadata_submission.json"

    if os.path.exists(metadata_path):
        with open(metadata_path, "r") as f:
            try:
                data = json.load(f)
            except json.JSONDecodeError:
                data = []
    else:
        data = []

    if 0 <= index < len(data):
        data[index] = {
            "images": [url.split("/")[-1] for url in payload.get("images", [])],
            "feature_desc": payload.get("feature_desc", ""),
            "generated_test_script": payload.get("generated_test_script", ""),
            "ground_truth_script": payload.get("ground_truth_script", "")
        }
        with open(metadata_path, "w") as f:
            json.dump(data, f, indent=2)

    return jsonify({"status": "ok"})


@app.route('/back', methods=['GET'])
def go_back():
    return redirect('/')


@app.route("/export-excel", methods=["GET"])
def export_excel():
    # 1) Đọc dataset
    metadata_path = os.path.join(UPLOAD_FOLDER, "metadata_submission.json")
    datasets = []
    if os.path.exists(metadata_path):
        try:
            with open(metadata_path, "r", encoding="utf-8") as f:
                datasets = json.load(f)
                if not isinstance(datasets, list):
                    datasets = []
        except json.JSONDecodeError:
            datasets = []

    # 2) Mở file template
    template_path = os.path.join(os.path.dirname(__file__), "report.xlsx")
    if not os.path.exists(template_path):
        return jsonify({"error": "Không tìm thấy report.xlsx"}), 404

    wb = load_workbook(template_path)
    ws1 = wb[wb.sheetnames[0]]
    ws2 = wb[wb.sheetnames[1]]

    # ===== Helpers: copy format (style, border, fill, number formats, alignment) từ 1 hàng mẫu =====

    def copy_row_format(ws, src_row: int, dst_row: int, max_col: int = None):
        """
        Clone style/format từ hàng src_row sang dst_row.
        Dùng copy(...) cho từng thuộc tính style để tránh StyleProxy error.
        """
        if max_col is None:
            max_col = ws.max_column

        # copy chiều cao hàng nếu có
        src_dim = ws.row_dimensions.get(src_row)
        if src_dim and src_dim.height is not None:
            ws.row_dimensions[dst_row].height = src_dim.height

        for col in range(1, max_col + 1):
            s = ws.cell(row=src_row, column=col)
            d = ws.cell(row=dst_row, column=col)

            if s.has_style:
                # Mỗi thuộc tính style cần copy riêng
                if s.font:
                    d.font = copy(s.font)
                if s.border:
                    d.border = copy(s.border)
                if s.fill:
                    d.fill = copy(s.fill)
                # number_format là chuỗi nên gán thẳng
                d.number_format = s.number_format
                if s.protection:
                    d.protection = copy(s.protection)
                if s.alignment:
                    d.alignment = copy(s.alignment)

    # =========================
    # SHEET 1: điền từ hàng 13, clone format hàng 13
    # =========================
    start_row = 13
    template_row_s1 = 13
    max_col_s1 = ws1.max_column

    def safe_unmerge(ws, range_str):
        for r in list(ws.merged_cells.ranges):
            if str(r) == range_str:
                ws.unmerge_cells(range_str)

    for i, item in enumerate(datasets, start=1):
        row = start_row + (i - 1)
        id_str = f"{i:04d}"

        # Clone format của hàng 13
        copy_row_format(ws1, template_row_s1, row, max_col_s1)

        # Ghi dữ liệu
        ws1[f"B{row}"] = id_str

        # E:F  -> IMG_LST_xxxx
        safe_unmerge(ws1, f"E{row}:F{row}")
        ws1.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        ws1[f"E{row}"] = f"IMG_LST_{id_str}"

        # G:H  -> feature_desc
        safe_unmerge(ws1, f"G{row}:H{row}")
        ws1.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
        ws1[f"G{row}"] = item.get("feature_desc", "")
        ws1[f"G{row}"].alignment = Alignment(wrap_text=True, vertical="top")

        # I:J  -> ground_truth_script
        safe_unmerge(ws1, f"I{row}:J{row}")
        ws1.merge_cells(start_row=row, start_column=9, end_row=row, end_column=10)
        ws1[f"I{row}"] = item.get("ground_truth_script", "")
        ws1[f"I{row}"].alignment = Alignment(wrap_text=True, vertical="top")

    # =========================
    # SHEET 2: ghi từ hàng 2
    # - A = IMG_LST_xxxx
    # - Ảnh: chèn từ cột B trở đi, kéo dài ngang (không xuống dòng)
    # - Hàng >=3 copy format của hàng 2
    # =========================
    template_row_s2 = 2
    max_col_s2 = max(ws2.max_column, 10)  # tạm thời
    # đảm bảo ít nhất các cột đầu có width khá ổn
    ws2.column_dimensions['A'].width = ws2.column_dimensions['A'].width or 18

    # chiều cao hàng mẫu 2 (nếu có)
    templ_h2 = ws2.row_dimensions[template_row_s2].height or 60

    for i, item in enumerate(datasets, start=1):
        row2 = 1 + i  # dataset 1 -> hàng 2, dataset 2 -> hàng 3,...
        id_str = f"{i:04d}"

        # Clone format của hàng 2 cho các hàng >=3
        if row2 >= 3:
            copy_row_format(ws2, template_row_s2, row2, max_col_s2)
        # Chiều cao hàng (để ảnh vừa vặn)
        ws2.row_dimensions[row2].height = templ_h2 or 60

        # A = IMG_LST_xxxx
        ws2[f"A{row2}"] = f"IMG_LST_{id_str}"

        # Ảnh: từ cột B trở đi, không xuống dòng
        imgs = item.get("images", []) or []
        thumb_w, thumb_h = 70, 70
        start_col = 2  # B

        for idx, filename in enumerate(imgs):
            img_path = os.path.join(UPLOAD_FOLDER, filename)
            if not os.path.exists(img_path):
                continue
            try:
                xlimg = XLImage(img_path)
                xlimg.width = thumb_w
                xlimg.height = thumb_h

                col_idx = start_col + idx  # B, C, D, E, F, ...
                col_letter = get_column_letter(col_idx)
                # đảm bảo chiều rộng cột cho ảnh
                if ws2.column_dimensions[col_letter].width is None or ws2.column_dimensions[col_letter].width < 15:
                    ws2.column_dimensions[col_letter].width = 20

                xlimg.anchor = f"{col_letter}{row2}"
                ws2.add_image(xlimg)
            except Exception:
                pass

    # 3) Lưu và trả file
    out_name = f"report_filled_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out_path = os.path.join(os.path.dirname(__file__), out_name)
    wb.save(out_path)
    return send_file(out_path, as_attachment=True, download_name=out_name)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001)
